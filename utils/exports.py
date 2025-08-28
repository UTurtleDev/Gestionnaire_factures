import csv
import io
import tempfile
import subprocess
import zipfile
from datetime import datetime
from django.http import HttpResponse
from django.conf import settings
from django.db import connection
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from clients.models import Client, Contact
from affaires.models import Affaire
from factures.models import Invoice, Payment


def export_database():
    """Export complet de la base de données SQLite"""
    try:
        with tempfile.NamedTemporaryFile(mode='w', suffix='.sql', delete=False) as temp_file:
            # Utiliser sqlite3 pour faire un dump de la base
            result = subprocess.run(
                ['sqlite3', settings.DATABASES['default']['NAME'], '.dump'],
                stdout=temp_file,
                stderr=subprocess.PIPE,
                text=True
            )
            
            if result.returncode != 0:
                raise Exception(f"Erreur lors de l'export: {result.stderr}")
            
            temp_file.flush()
            
            # Lire le fichier et le retourner comme réponse
            with open(temp_file.name, 'r') as f:
                sql_content = f.read()
            
            response = HttpResponse(sql_content, content_type='application/sql')
            response['Content-Disposition'] = f'attachment; filename="backup_database_{datetime.now().strftime("%Y%m%d_%H%M%S")}.sql"'
            return response
            
    except Exception as e:
        raise Exception(f"Erreur lors de l'export de la base de données: {str(e)}")


def export_clients_csv():
    """Export des clients au format CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="clients_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # BOM pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'Nom de l\'entité', 'Adresse', 'Code postal', 'Ville', 
        'Contact', 'Téléphone', 'Email', 'Total affaires (€)'
    ])
    
    clients = Client.objects.all().order_by('entity_name')
    for client in clients:
        writer.writerow([
            client.entity_name or '',
            client.address or '',
            client.zip_code or '',
            client.city or '',
            client.contact or '',
            client.phone_number or '',
            client.email or '',
            f"{client.total_affaire_client:.2f}".replace('.', ',')
        ])
    
    return response


def export_clients_xlsx():
    """Export des clients au format Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    
    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7FAEDC", end_color="7FAEDC", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    
    # En-têtes
    headers = [
        'Nom de l\'entité', 'Adresse', 'Code postal', 'Ville', 
        'Contact', 'Téléphone', 'Email', 'Total affaires (€)'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Données
    clients = Client.objects.all().order_by('entity_name')
    for row, client in enumerate(clients, 2):
        ws.cell(row=row, column=1, value=client.entity_name or '')
        ws.cell(row=row, column=2, value=client.address or '')
        ws.cell(row=row, column=3, value=client.zip_code or '')
        ws.cell(row=row, column=4, value=client.city or '')
        ws.cell(row=row, column=5, value=client.contact or '')
        ws.cell(row=row, column=6, value=client.phone_number or '')
        ws.cell(row=row, column=7, value=client.email or '')
        ws.cell(row=row, column=8, value=float(client.total_affaire_client))
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse HTTP
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="clients_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response


def export_affaires_csv(date_debut=None, date_fin=None):
    """Export des affaires au format CSV avec filtrage par date"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="affaires_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # BOM pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'Numéro affaire', 'Client', 'Description', 'Budget (€)', 
        'Total facturé HT (€)', 'Reste à facturer (€)', 'Taux d\'avancement (%)',
        'Contact principal', 'Auteur'
    ])
    
    affaires = Affaire.objects.all().order_by('affaire_number')
    
    # Filtrage par date si spécifié (basé sur la date de création des factures associées)
    if date_debut or date_fin:
        affaire_ids = set()
        invoices = Invoice.objects.all()
        if date_debut:
            invoices = invoices.filter(date__gte=date_debut)
        if date_fin:
            invoices = invoices.filter(date__lte=date_fin)
        affaire_ids.update(invoices.values_list('affaire_id', flat=True))
        affaires = affaires.filter(id__in=affaire_ids)
    
    for affaire in affaires:
        contact_principal = affaire.contact_principal
        contact_nom = str(contact_principal) if contact_principal else ''
        
        author_name = ''
        if affaire.author:
            author_name = f"{affaire.author.first_name} {affaire.author.last_name}".strip()
            if not author_name:
                author_name = affaire.author.email
        
        writer.writerow([
            affaire.affaire_number,
            affaire.client_entity_name or '',
            affaire.affaire_description or '',
            f"{affaire.budget:.2f}".replace('.', ','),
            f"{affaire.total_facture_ht:.2f}".replace('.', ','),
            f"{affaire.reste_a_facturer:.2f}".replace('.', ','),
            f"{affaire.taux_avancement:.1f}".replace('.', ','),
            contact_nom,
            author_name
        ])
    
    return response


def export_affaires_xlsx(date_debut=None, date_fin=None):
    """Export des affaires au format Excel avec filtrage par date"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Affaires"
    
    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7FAEDC", end_color="7FAEDC", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    
    # En-têtes
    headers = [
        'Numéro affaire', 'Client', 'Description', 'Budget (€)', 
        'Total facturé HT (€)', 'Reste à facturer (€)', 'Taux d\'avancement (%)',
        'Contact principal', 'Auteur'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Données
    affaires = Affaire.objects.all().order_by('affaire_number')
    
    # Filtrage par date si spécifié
    if date_debut or date_fin:
        affaire_ids = set()
        invoices = Invoice.objects.all()
        if date_debut:
            invoices = invoices.filter(date__gte=date_debut)
        if date_fin:
            invoices = invoices.filter(date__lte=date_fin)
        affaire_ids.update(invoices.values_list('affaire_id', flat=True))
        affaires = affaires.filter(id__in=affaire_ids)
    
    for row, affaire in enumerate(affaires, 2):
        contact_principal = affaire.contact_principal
        contact_nom = str(contact_principal) if contact_principal else ''
        
        author_name = ''
        if affaire.author:
            author_name = f"{affaire.author.first_name} {affaire.author.last_name}".strip()
            if not author_name:
                author_name = affaire.author.email
        
        ws.cell(row=row, column=1, value=affaire.affaire_number)
        ws.cell(row=row, column=2, value=affaire.client_entity_name or '')
        ws.cell(row=row, column=3, value=affaire.affaire_description or '')
        ws.cell(row=row, column=4, value=float(affaire.budget))
        ws.cell(row=row, column=5, value=float(affaire.total_facture_ht))
        ws.cell(row=row, column=6, value=float(affaire.reste_a_facturer))
        ws.cell(row=row, column=7, value=float(affaire.taux_avancement))
        ws.cell(row=row, column=8, value=contact_nom)
        ws.cell(row=row, column=9, value=author_name)
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse HTTP
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="affaires_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response


def export_factures_csv(date_debut=None, date_fin=None):
    """Export des factures au format CSV avec filtrage par date"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="factures_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # BOM pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'Numéro facture', 'Date', 'Type', 'Client', 'Affaire', 'Objet', 
        'Montant HT (€)', 'Taux TVA (%)', 'Montant TTC (€)', 'Statut',
        'Date échéance', 'Solde (€)', 'Contact', 'Auteur'
    ])
    
    factures = Invoice.objects.all().order_by('-date')
    
    # Filtrage par date
    if date_debut:
        factures = factures.filter(date__gte=date_debut)
    if date_fin:
        factures = factures.filter(date__lte=date_fin)
    
    for facture in factures:
        contact_nom = str(facture.contact) if facture.contact else ''
        
        author_name = ''
        if facture.author:
            author_name = f"{facture.author.first_name} {facture.author.last_name}".strip()
            if not author_name:
                author_name = facture.author.email
        
        writer.writerow([
            facture.invoice_number,
            facture.date.strftime('%d/%m/%Y'),
            facture.get_type_display(),
            facture.client_entity_name or '',
            facture.affaire.affaire_number if facture.affaire else '',
            facture.invoice_object or '',
            f"{facture.amount_ht:.2f}".replace('.', ','),
            f"{facture.vat_rate:.1f}".replace('.', ','),
            f"{facture.amount_ttc:.2f}".replace('.', ','),
            facture.get_statut_display(),
            facture.due_date.strftime('%d/%m/%Y'),
            f"{facture.balance:.2f}".replace('.', ','),
            contact_nom,
            author_name
        ])
    
    return response


def export_factures_xlsx(date_debut=None, date_fin=None):
    """Export des factures au format Excel avec filtrage par date"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Factures"
    
    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7FAEDC", end_color="7FAEDC", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    
    # En-têtes
    headers = [
        'Numéro facture', 'Date', 'Type', 'Client', 'Affaire', 'Objet', 
        'Montant HT (€)', 'Taux TVA (%)', 'Montant TTC (€)', 'Statut',
        'Date échéance', 'Solde (€)', 'Contact', 'Auteur'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Données
    factures = Invoice.objects.all().order_by('-date')
    
    # Filtrage par date
    if date_debut:
        factures = factures.filter(date__gte=date_debut)
    if date_fin:
        factures = factures.filter(date__lte=date_fin)
    
    for row, facture in enumerate(factures, 2):
        contact_nom = str(facture.contact) if facture.contact else ''
        
        author_name = ''
        if facture.author:
            author_name = f"{facture.author.first_name} {facture.author.last_name}".strip()
            if not author_name:
                author_name = facture.author.email
        
        ws.cell(row=row, column=1, value=facture.invoice_number)
        ws.cell(row=row, column=2, value=facture.date)
        ws.cell(row=row, column=3, value=facture.get_type_display())
        ws.cell(row=row, column=4, value=facture.client_entity_name or '')
        ws.cell(row=row, column=5, value=facture.affaire.affaire_number if facture.affaire else '')
        ws.cell(row=row, column=6, value=facture.invoice_object or '')
        ws.cell(row=row, column=7, value=float(facture.amount_ht))
        ws.cell(row=row, column=8, value=float(facture.vat_rate))
        ws.cell(row=row, column=9, value=float(facture.amount_ttc))
        ws.cell(row=row, column=10, value=facture.get_statut_display())
        ws.cell(row=row, column=11, value=facture.due_date)
        ws.cell(row=row, column=12, value=float(facture.balance))
        ws.cell(row=row, column=13, value=contact_nom)
        ws.cell(row=row, column=14, value=author_name)
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse HTTP
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="factures_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response


def export_database_csv():
    """Export complet de la base de données au format CSV (toutes les tables)"""
    output = io.BytesIO()
    
    # Créer un fichier ZIP contenant tous les CSV
    
    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zipf:
        
        # Export des clients
        clients_csv = io.StringIO()
        clients_csv.write('\ufeff')  # BOM pour Excel
        writer = csv.writer(clients_csv, delimiter=';')
        writer.writerow([
            'Nom de l\'entité', 'Adresse', 'Code postal', 'Ville', 
            'Contact', 'Téléphone', 'Email', 'Total affaires (€)'
        ])
        
        clients = Client.objects.all().order_by('entity_name')
        for client in clients:
            writer.writerow([
                client.entity_name or '',
                client.address or '',
                client.zip_code or '',
                client.city or '',
                client.contact or '',
                client.phone_number or '',
                client.email or '',
                f"{client.total_affaire_client:.2f}".replace('.', ',')
            ])
        
        zipf.writestr('clients.csv', clients_csv.getvalue().encode('utf-8'))
        
        # Export des contacts
        contacts_csv = io.StringIO()
        contacts_csv.write('\ufeff')
        writer = csv.writer(contacts_csv, delimiter=';')
        writer.writerow([
            'Nom', 'Prénom', 'Fonction', 'Téléphone', 'Email', 
            'Principal', 'Affaire', 'Client'
        ])
        
        contacts = Contact.objects.all().order_by('nom')
        for contact in contacts:
            writer.writerow([
                contact.nom or '',
                contact.prenom or '',
                contact.fonction or '',
                contact.phone_number or '',
                contact.email or '',
                'Oui' if contact.is_principal else 'Non',
                contact.affaire.affaire_number if contact.affaire else '',
                contact.affaire.client_entity_name if contact.affaire else ''
            ])
        
        zipf.writestr('contacts.csv', contacts_csv.getvalue().encode('utf-8'))
        
        # Export des affaires
        affaires_csv = io.StringIO()
        affaires_csv.write('\ufeff')
        writer = csv.writer(affaires_csv, delimiter=';')
        writer.writerow([
            'Numéro affaire', 'Client', 'Description', 'Budget (€)', 
            'Total facturé HT (€)', 'Reste à facturer (€)', 'Taux d\'avancement (%)',
            'Contact principal', 'Auteur'
        ])
        
        affaires = Affaire.objects.all().order_by('affaire_number')
        for affaire in affaires:
            contact_principal = affaire.contact_principal
            contact_nom = str(contact_principal) if contact_principal else ''
            
            author_name = ''
            if affaire.author:
                author_name = f"{affaire.author.first_name} {affaire.author.last_name}".strip()
                if not author_name:
                    author_name = affaire.author.email
            
            writer.writerow([
                affaire.affaire_number,
                affaire.client_entity_name or '',
                affaire.affaire_description or '',
                f"{affaire.budget:.2f}".replace('.', ','),
                f"{affaire.total_facture_ht:.2f}".replace('.', ','),
                f"{affaire.reste_a_facturer:.2f}".replace('.', ','),
                f"{affaire.taux_avancement:.1f}".replace('.', ','),
                contact_nom,
                author_name
            ])
        
        zipf.writestr('affaires.csv', affaires_csv.getvalue().encode('utf-8'))
        
        # Export des factures
        factures_csv = io.StringIO()
        factures_csv.write('\ufeff')
        writer = csv.writer(factures_csv, delimiter=';')
        writer.writerow([
            'Numéro facture', 'Date', 'Type', 'Client', 'Affaire', 'Objet', 
            'Montant HT (€)', 'Taux TVA (%)', 'Montant TTC (€)', 'Statut',
            'Date échéance', 'Solde (€)', 'Contact', 'Auteur'
        ])
        
        factures = Invoice.objects.all().order_by('-date')
        for facture in factures:
            contact_nom = str(facture.contact) if facture.contact else ''
            
            author_name = ''
            if facture.author:
                author_name = f"{facture.author.first_name} {facture.author.last_name}".strip()
                if not author_name:
                    author_name = facture.author.email
            
            writer.writerow([
                facture.invoice_number,
                facture.date.strftime('%d/%m/%Y'),
                facture.get_type_display(),
                facture.client_entity_name or '',
                facture.affaire.affaire_number if facture.affaire else '',
                facture.invoice_object or '',
                f"{facture.amount_ht:.2f}".replace('.', ','),
                f"{facture.vat_rate:.1f}".replace('.', ','),
                f"{facture.amount_ttc:.2f}".replace('.', ','),
                facture.get_statut_display(),
                facture.due_date.strftime('%d/%m/%Y'),
                f"{facture.balance:.2f}".replace('.', ','),
                contact_nom,
                author_name
            ])
        
        zipf.writestr('factures.csv', factures_csv.getvalue().encode('utf-8'))
        
        # Export des paiements
        paiements_csv = io.StringIO()
        paiements_csv.write('\ufeff')
        writer = csv.writer(paiements_csv, delimiter=';')
        writer.writerow([
            'Date', 'Montant (€)', 'Numéro facture', 'Moyen de paiement'
        ])
        
        paiements = Payment.objects.all().order_by('-date')
        for paiement in paiements:
            writer.writerow([
                paiement.date.strftime('%d/%m/%Y'),
                f"{paiement.amount:.2f}".replace('.', ','),
                paiement.invoice.invoice_number,
                paiement.payment_method or ''
            ])
        
        zipf.writestr('paiements.csv', paiements_csv.getvalue().encode('utf-8'))
    
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/zip'
    )
    response['Content-Disposition'] = f'attachment; filename="export_base_donnees_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip"'
    return response


def export_database_xlsx():
    """Export complet de la base de données au format Excel (toutes les tables)"""
    wb = Workbook()
    
    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7FAEDC", end_color="7FAEDC", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    
    # Supprimer la feuille par défaut
    wb.remove(wb.active)
    
    # Feuille Clients
    ws_clients = wb.create_sheet("Clients")
    headers_clients = [
        'Nom de l\'entité', 'Adresse', 'Code postal', 'Ville', 
        'Contact', 'Téléphone', 'Email', 'Total affaires (€)'
    ]
    
    for col, header in enumerate(headers_clients, 1):
        cell = ws_clients.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    clients = Client.objects.all().order_by('entity_name')
    for row, client in enumerate(clients, 2):
        ws_clients.cell(row=row, column=1, value=client.entity_name or '')
        ws_clients.cell(row=row, column=2, value=client.address or '')
        ws_clients.cell(row=row, column=3, value=client.zip_code or '')
        ws_clients.cell(row=row, column=4, value=client.city or '')
        ws_clients.cell(row=row, column=5, value=client.contact or '')
        ws_clients.cell(row=row, column=6, value=client.phone_number or '')
        ws_clients.cell(row=row, column=7, value=client.email or '')
        ws_clients.cell(row=row, column=8, value=float(client.total_affaire_client))
    
    # Feuille Contacts
    ws_contacts = wb.create_sheet("Contacts")
    headers_contacts = [
        'Nom', 'Prénom', 'Fonction', 'Téléphone', 'Email', 
        'Principal', 'Affaire', 'Client'
    ]
    
    for col, header in enumerate(headers_contacts, 1):
        cell = ws_contacts.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    contacts = Contact.objects.all().order_by('nom')
    for row, contact in enumerate(contacts, 2):
        ws_contacts.cell(row=row, column=1, value=contact.nom or '')
        ws_contacts.cell(row=row, column=2, value=contact.prenom or '')
        ws_contacts.cell(row=row, column=3, value=contact.fonction or '')
        ws_contacts.cell(row=row, column=4, value=contact.phone_number or '')
        ws_contacts.cell(row=row, column=5, value=contact.email or '')
        ws_contacts.cell(row=row, column=6, value='Oui' if contact.is_principal else 'Non')
        ws_contacts.cell(row=row, column=7, value=contact.affaire.affaire_number if contact.affaire else '')
        ws_contacts.cell(row=row, column=8, value=contact.affaire.client_entity_name if contact.affaire else '')
    
    # Feuille Affaires
    ws_affaires = wb.create_sheet("Affaires")
    headers_affaires = [
        'Numéro affaire', 'Client', 'Description', 'Budget (€)', 
        'Total facturé HT (€)', 'Reste à facturer (€)', 'Taux d\'avancement (%)',
        'Contact principal', 'Auteur'
    ]
    
    for col, header in enumerate(headers_affaires, 1):
        cell = ws_affaires.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    affaires = Affaire.objects.all().order_by('affaire_number')
    for row, affaire in enumerate(affaires, 2):
        contact_principal = affaire.contact_principal
        contact_nom = str(contact_principal) if contact_principal else ''
        
        author_name = ''
        if affaire.author:
            author_name = f"{affaire.author.first_name} {affaire.author.last_name}".strip()
            if not author_name:
                author_name = affaire.author.email
        
        ws_affaires.cell(row=row, column=1, value=affaire.affaire_number)
        ws_affaires.cell(row=row, column=2, value=affaire.client_entity_name or '')
        ws_affaires.cell(row=row, column=3, value=affaire.affaire_description or '')
        ws_affaires.cell(row=row, column=4, value=float(affaire.budget))
        ws_affaires.cell(row=row, column=5, value=float(affaire.total_facture_ht))
        ws_affaires.cell(row=row, column=6, value=float(affaire.reste_a_facturer))
        ws_affaires.cell(row=row, column=7, value=float(affaire.taux_avancement))
        ws_affaires.cell(row=row, column=8, value=contact_nom)
        ws_affaires.cell(row=row, column=9, value=author_name)
    
    # Feuille Factures
    ws_factures = wb.create_sheet("Factures")
    headers_factures = [
        'Numéro facture', 'Date', 'Type', 'Client', 'Affaire', 'Objet', 
        'Montant HT (€)', 'Taux TVA (%)', 'Montant TTC (€)', 'Statut',
        'Date échéance', 'Solde (€)', 'Contact', 'Auteur'
    ]
    
    for col, header in enumerate(headers_factures, 1):
        cell = ws_factures.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    factures = Invoice.objects.all().order_by('-date')
    for row, facture in enumerate(factures, 2):
        contact_nom = str(facture.contact) if facture.contact else ''
        
        author_name = ''
        if facture.author:
            author_name = f"{facture.author.first_name} {facture.author.last_name}".strip()
            if not author_name:
                author_name = facture.author.email
        
        ws_factures.cell(row=row, column=1, value=facture.invoice_number)
        ws_factures.cell(row=row, column=2, value=facture.date)
        ws_factures.cell(row=row, column=3, value=facture.get_type_display())
        ws_factures.cell(row=row, column=4, value=facture.client_entity_name or '')
        ws_factures.cell(row=row, column=5, value=facture.affaire.affaire_number if facture.affaire else '')
        ws_factures.cell(row=row, column=6, value=facture.invoice_object or '')
        ws_factures.cell(row=row, column=7, value=float(facture.amount_ht))
        ws_factures.cell(row=row, column=8, value=float(facture.vat_rate))
        ws_factures.cell(row=row, column=9, value=float(facture.amount_ttc))
        ws_factures.cell(row=row, column=10, value=facture.get_statut_display())
        ws_factures.cell(row=row, column=11, value=facture.due_date)
        ws_factures.cell(row=row, column=12, value=float(facture.balance))
        ws_factures.cell(row=row, column=13, value=contact_nom)
        ws_factures.cell(row=row, column=14, value=author_name)
    
    # Feuille Paiements
    ws_paiements = wb.create_sheet("Paiements")
    headers_paiements = [
        'Date', 'Montant (€)', 'Numéro facture', 'Moyen de paiement'
    ]
    
    for col, header in enumerate(headers_paiements, 1):
        cell = ws_paiements.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    paiements = Payment.objects.all().order_by('-date')
    for row, paiement in enumerate(paiements, 2):
        ws_paiements.cell(row=row, column=1, value=paiement.date)
        ws_paiements.cell(row=row, column=2, value=float(paiement.amount))
        ws_paiements.cell(row=row, column=3, value=paiement.invoice.invoice_number)
        ws_paiements.cell(row=row, column=4, value=paiement.payment_method or '')
    
    # Ajuster la largeur des colonnes pour toutes les feuilles
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse HTTP
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="export_base_donnees_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response


def export_contacts_csv():
    """Export des contacts au format CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="contacts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # BOM pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'Nom', 'Prénom', 'Fonction', 'Téléphone', 'Email', 
        'Principal', 'Numéro affaire', 'Client'
    ])
    
    contacts = Contact.objects.all().order_by('nom')
    for contact in contacts:
        writer.writerow([
            contact.nom or '',
            contact.prenom or '',
            contact.fonction or '',
            contact.phone_number or '',
            contact.email or '',
            'Oui' if contact.is_principal else 'Non',
            contact.affaire.affaire_number if contact.affaire else '',
            contact.affaire.client_entity_name if contact.affaire else ''
        ])
    
    return response


def export_contacts_xlsx():
    """Export des contacts au format Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    
    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7FAEDC", end_color="7FAEDC", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    
    # En-têtes
    headers = [
        'Nom', 'Prénom', 'Fonction', 'Téléphone', 'Email', 
        'Principal', 'Numéro affaire', 'Client'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Données
    contacts = Contact.objects.all().order_by('nom')
    for row, contact in enumerate(contacts, 2):
        ws.cell(row=row, column=1, value=contact.nom or '')
        ws.cell(row=row, column=2, value=contact.prenom or '')
        ws.cell(row=row, column=3, value=contact.fonction or '')
        ws.cell(row=row, column=4, value=contact.phone_number or '')
        ws.cell(row=row, column=5, value=contact.email or '')
        ws.cell(row=row, column=6, value='Oui' if contact.is_principal else 'Non')
        ws.cell(row=row, column=7, value=contact.affaire.affaire_number if contact.affaire else '')
        ws.cell(row=row, column=8, value=contact.affaire.client_entity_name if contact.affaire else '')
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse HTTP
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="contacts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response


def export_reglements_csv(date_debut=None, date_fin=None):
    """Export des règlements au format CSV avec filtrage par date"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="reglements_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # BOM pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'Date', 'Montant (€)', 'Numéro facture', 'Client', 'Affaire', 'Moyen de paiement'
    ])
    
    paiements = Payment.objects.all().order_by('-date')
    
    # Filtrage par date
    if date_debut:
        paiements = paiements.filter(date__gte=date_debut)
    if date_fin:
        paiements = paiements.filter(date__lte=date_fin)
    
    for paiement in paiements:
        writer.writerow([
            paiement.date.strftime('%d/%m/%Y'),
            f"{paiement.amount:.2f}".replace('.', ','),
            paiement.invoice.invoice_number,
            paiement.invoice.client_entity_name or '',
            paiement.invoice.affaire.affaire_number if paiement.invoice.affaire else '',
            paiement.payment_method or ''
        ])
    
    return response


def export_reglements_xlsx(date_debut=None, date_fin=None):
    """Export des règlements au format Excel avec filtrage par date"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Règlements"
    
    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7FAEDC", end_color="7FAEDC", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    
    # En-têtes
    headers = [
        'Date', 'Montant (€)', 'Numéro facture', 'Client', 'Affaire', 'Moyen de paiement'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Données
    paiements = Payment.objects.all().order_by('-date')
    
    # Filtrage par date
    if date_debut:
        paiements = paiements.filter(date__gte=date_debut)
    if date_fin:
        paiements = paiements.filter(date__lte=date_fin)
    
    for row, paiement in enumerate(paiements, 2):
        ws.cell(row=row, column=1, value=paiement.date)
        ws.cell(row=row, column=2, value=float(paiement.amount))
        ws.cell(row=row, column=3, value=paiement.invoice.invoice_number)
        ws.cell(row=row, column=4, value=paiement.invoice.client_entity_name or '')
        ws.cell(row=row, column=5, value=paiement.invoice.affaire.affaire_number if paiement.invoice.affaire else '')
        ws.cell(row=row, column=6, value=paiement.payment_method or '')
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse HTTP
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reglements_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response